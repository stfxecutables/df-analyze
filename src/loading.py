from __future__ import annotations

# fmt: off
import sys  # isort: skip
from pathlib import Path  # isort: skip
ROOT = Path(__file__).resolve().parent.parent  # isort: skip
sys.path.append(str(ROOT))  # isort: skip
# fmt: on

from io import StringIO
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame

from src._constants import SIMPLE_CSV, SIMPLE_ODS, SIMPLE_XLSX


def load_excel(path: Path) -> tuple[DataFrame, dict[int, str]]:
    wb = load_workbook(path, data_only=True)
    sheetnames = wb.sheetnames
    if len(sheetnames) != 1:
        raise RuntimeError(
            "Found multiple sheets in Excel workbook. Make sure the Excel file "
            "has only a single sheet formatted correctly for df-analyze."
        )
    ws: Worksheet = wb[sheetnames[0]]
    row: tuple[Any]
    meta = {}
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):  # type: ignore
        first = row[0]
        if first is not None and str(first).strip().startswith("--"):
            line = " ".join([str(value) for value in row if value is not None])
            meta[i] = line
            continue
        if all(val is None for val in row):  # ignore blank rows
            continue
        data.append(row)

    df = DataFrame(data=data[1:], columns=data[0])

    return df, meta


def load_csv(path: Path, separator: str = ",") -> tuple[DataFrame, dict[int, str]]:
    with open(path, "r") as handle:
        lines = handle.readlines()

    meta = {}
    header = None
    empties = ["\n", ""]
    for i, line in enumerate(lines):
        if line in empties or (line.replace(separator, "") in empties):
            continue  # ignore blanks
        if line.startswith("--"):
            meta[i] = line.replace(separator, " ").strip()
        else:
            header = i
            break

    if header is None:
        raise RuntimeError(
            f"Could not find start of data (column names) in csv file at "
            f"{path}. The file is likely incorrectly formatted."
        )

    data = StringIO("".join(lines[header:]))
    df = pd.read_csv(data, sep=separator)
    return df, meta


if __name__ == "__main__":
    df, meta = load_excel(SIMPLE_XLSX)
    df_meta = load_csv(SIMPLE_CSV)
    print(type(list(df.dtypes)[0]))
